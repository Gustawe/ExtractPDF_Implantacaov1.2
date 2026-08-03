from __future__ import annotations

from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from folha_pdf_xlsx.layout_writer import write_layout_workbook
from folha_pdf_xlsx.models import (
    DocumentMetadata,
    EmployeeRecord,
    PayrollDocument,
    PayrollEvent,
    ProcessingIssue,
)
from folha_pdf_xlsx.validation import validate_document
from folha_pdf_xlsx.writer import write_workbook


def _employee(*, events: bool, event_earnings: str = "100.00") -> EmployeeRecord:
    employee = EmployeeRecord(
        source_file="folha.pdf",
        page=2,
        employee_key="folha.pdf|2|EMPREGADO|71|1",
        record_type="EMPREGADO",
        registration="71",
        name="Pessoa de Teste",
        total_earnings=Decimal("100.00"),
        total_discounts=Decimal("10.00"),
        net_amount=Decimal("90.00"),
    )
    if events:
        employee.events.extend(
            (
                PayrollEvent(
                    source_file="folha.pdf",
                    page=2,
                    employee_key=employee.employee_key,
                    registration="71",
                    code="1",
                    description="PROVENTO",
                    reference="",
                    value=Decimal(event_earnings),
                    kind="P",
                ),
                PayrollEvent(
                    source_file="folha.pdf",
                    page=2,
                    employee_key=employee.employee_key,
                    registration="71",
                    code="2",
                    description="DESCONTO",
                    reference="",
                    value=Decimal("10.00"),
                    kind="D",
                ),
            )
        )
    return employee


def _document(employee: EmployeeRecord, calculation: str = "Folha mensal") -> PayrollDocument:
    return PayrollDocument(
        source_path=Path("folha.pdf"),
        metadata=DocumentMetadata(
            source_file="folha.pdf",
            calculation=calculation,
            competence="11/2025",
            layout_profile="extrato_mensal_v1",
        ),
        employees=[employee],
    )


def test_recognized_thirteenth_without_events_is_not_applicable(
    tmp_path: Path,
) -> None:
    document = _document(
        _employee(events=False),
        calculation="Adiantamento de 13º Salário — 1ª Parcela",
    )
    document.issues.append(
        ProcessingIssue(
            source_file="folha.pdf",
            severity="AVISO",
            code="RESUMO_FISCAL_AUSENTE",
            message="Página de resumo fiscal não encontrada.",
        )
    )

    document.validations = validate_document(document)

    event_checks = document.validations[:2]
    assert [check.status for check in event_checks] == [
        "NÃO APLICÁVEL",
        "NÃO APLICÁVEL",
    ]
    assert all("13º" in check.message for check in event_checks)
    assert not document.details.divergence_count
    assert document.status == "APROVADO COM AVISOS"
    assert [issue.code for issue in document.issues] == ["RESUMO_FISCAL_AUSENTE"]

    output = tmp_path / "decimo-terceiro.xlsx"
    write_layout_workbook(document, output)
    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == ["Folha", "Validações", "Ocorrências"]
    assert workbook["Validações"]["A2"].value == "NÃO APLICÁVEL"
    assert "[Red]" not in workbook["Validações"]["G2"].number_format
    target = workbook["Folha"][event_checks[0].target_cell.split("!", 1)[1]]
    assert target.fill.fgColor.rgb != "00FFF2CC"
    assert target.comment is None
    workbook.close()


def test_normal_layout_without_events_records_a_reading_warning() -> None:
    document = _document(_employee(events=False))

    document.validations = validate_document(document)

    assert [check.status for check in document.validations[:2]] == ["AVISO", "AVISO"]
    assert document.issues[0].code == "EVENTOS_DETALHADOS_NAO_LIDOS"
    assert document.issues[0].employee_name == "Pessoa de Teste"
    assert document.status == "APROVADO COM AVISOS"


def test_visual_writer_adds_audit_and_marks_mapped_divergence(tmp_path: Path) -> None:
    document = _document(_employee(events=True, event_earnings="90.00"))
    document.validations = validate_document(document)
    output = tmp_path / "visual.xlsx"

    write_layout_workbook(document, output)

    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == ["Folha", "Validações"]
    divergence = next(
        check for check in document.validations if check.status == "DIVERGÊNCIA"
    )
    assert divergence.target_cell.startswith("Folha!E")
    target = workbook["Folha"][divergence.target_cell.split("!", 1)[1]]
    assert target.fill.fgColor.rgb == "00FFF2CC"
    assert target.comment is not None
    assert "Esperado: R$ 100,00" in target.comment.text
    assert workbook["Validações"]["A2"].value == "DIVERGÊNCIA"
    workbook.close()


def test_generic_writer_uses_the_same_audit_structure(tmp_path: Path) -> None:
    document = _document(_employee(events=True, event_earnings="90.00"))
    document.validations = validate_document(document)
    document.issues.append(
        ProcessingIssue(
            source_file="folha.pdf",
            severity="AVISO",
            code="TESTE",
            message="Ocorrência de teste",
        )
    )
    output = tmp_path / "generico.xlsx"

    write_workbook([document], output)

    workbook = load_workbook(output, data_only=False)
    assert "Validações" in workbook.sheetnames
    assert "Ocorrências" in workbook.sheetnames
    divergence = next(
        check for check in document.validations if check.status == "DIVERGÊNCIA"
    )
    target = workbook["Funcionarios"][divergence.target_cell.split("!", 1)[1]]
    assert target.fill.fgColor.rgb == "00FFF2CC"
    assert target.comment is not None
    workbook.close()


def test_consistent_monthly_sheet_has_no_audit_tab_or_highlight(tmp_path: Path) -> None:
    document = _document(_employee(events=True))
    document.validations = validate_document(document)
    output = tmp_path / "consistente.xlsx"

    write_layout_workbook(document, output)

    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == ["Folha"]
    target = workbook["Folha"][document.validations[0].target_cell.split("!", 1)[1]]
    assert target.fill.fill_type != "solid" or target.fill.fgColor.rgb != "00FFF2CC"
    assert target.comment is None
    workbook.close()
