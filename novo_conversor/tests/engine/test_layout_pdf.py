from __future__ import annotations

import os
import tempfile
import unittest
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from folha_pdf_xlsx.layout_extractor import SystemPayrollLayoutExtractor
from folha_pdf_xlsx.layout_writer import write_layout_workbook
from folha_pdf_xlsx.models import (
    DocumentMetadata,
    EmployeeRecord,
    PayrollDocument,
    PayrollEvent,
)


SAMPLE = os.environ.get("FOLHA_LAYOUT_SAMPLE_PDF")


@unittest.skipUnless(SAMPLE and Path(SAMPLE).is_file(), "PDF de layout não definido")
class PayrollLayoutTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.document = SystemPayrollLayoutExtractor().extract(Path(SAMPLE))

    def test_reference_structure_is_recognized(self) -> None:
        self.assertEqual(self.document.page_count, 91)
        self.assertEqual(len(self.document.sections), 25)
        self.assertEqual(self.document.employee_count, 257)

    def test_page_continuations_stay_inside_employee_blocks(self) -> None:
        continued = [
            block
            for section in self.document.sections
            for block in section.employee_blocks
            if len({line.page for line in block}) > 1
        ]
        self.assertTrue(continued)
        for block in continued:
            self.assertTrue(any(line.text.startswith("Proventos:") for line in block))

    def test_writer_uses_one_sheet_and_typed_values(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "folha.xlsx"
            write_layout_workbook(self.document, output)
            workbook = load_workbook(output, data_only=False)
            self.assertEqual(workbook.sheetnames, ["Folha"])
            sheet = workbook["Folha"]
            self.assertEqual(
                sum(1 for row in sheet.iter_rows() if row[0].value == "Cód:"),
                257,
            )
            self.assertTrue(
                any(
                    isinstance(cell.value, datetime)
                    for row in sheet.iter_rows()
                    for cell in row
                )
            )
            self.assertTrue(
                any(
                    isinstance(cell.value, float)
                    for row in sheet.iter_rows()
                    for cell in row
                )
            )
            workbook.close()


class StructuredPayrollVisualWriterTests(unittest.TestCase):
    def test_structured_document_uses_approved_single_sheet_layout(self) -> None:
        employee = EmployeeRecord(
            source_file="folha.pdf",
            page=1,
            employee_key="folha.pdf|1|EMPREGADO|71|1",
            record_type="EMPREGADO",
            registration="71",
            name="Pessoa de Teste",
            status="Trabalhando",
            job_title="ANALISTA",
            salary=Decimal("4400.00"),
            total_earnings=Decimal("4500.00"),
            total_discounts=Decimal("740.00"),
            net_amount=Decimal("3760.00"),
            inss_base=Decimal("4500.00"),
            fgts_base=Decimal("4500.00"),
            fgts_value=Decimal("360.00"),
            irrf_base=Decimal("3900.00"),
        )
        employee.events = [
            PayrollEvent(
                source_file="folha.pdf",
                page=1,
                employee_key=employee.employee_key,
                registration="71",
                code="1",
                description="SALÁRIO",
                reference="30,00",
                value=Decimal("4500.00"),
                kind="P",
            ),
            PayrollEvent(
                source_file="folha.pdf",
                page=1,
                employee_key=employee.employee_key,
                registration="71",
                code="11",
                description="INSS",
                reference="11,00",
                value=Decimal("500.00"),
                kind="D",
            ),
            PayrollEvent(
                source_file="folha.pdf",
                page=1,
                employee_key=employee.employee_key,
                registration="71",
                code="48",
                description="VALE TRANSPORTE",
                reference="6,00",
                value=Decimal("240.00"),
                kind="D",
            ),
        ]
        document = PayrollDocument(
            source_path=Path("folha.pdf"),
            metadata=DocumentMetadata(
                source_file="folha.pdf",
                company_code="1035",
                company_name="EMPRESA TESTE",
                cnpj="00.000.000/0001-00",
                competence="01/2025",
            ),
            employees=[employee],
        )

        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "folha.xlsx"
            write_layout_workbook(document, output)
            workbook = load_workbook(output, data_only=True)
            self.assertEqual(workbook.sheetnames, ["Folha"])
            sheet = workbook["Folha"]
            self.assertEqual(sheet["A1"].value, "Folha de Pagamento — 01/01/2025 a 31/01/2025")
            self.assertEqual(sheet["A6"].value, "Cód:")
            self.assertEqual(sheet["B6"].value, 71)
            self.assertEqual(sheet["E6"].value, "Pessoa de Teste")
            self.assertEqual(sheet["D7"].value, "Situação:")
            self.assertEqual(sheet["E7"].value, "Trabalhando")
            self.assertEqual(sheet["G8"].value, 4500)
            self.assertEqual(sheet["H8"].value, 11)
            self.assertEqual(sheet["M8"].value, 500)
            self.assertIsNone(sheet["B9"].value)
            self.assertEqual(sheet["H9"].value, 48)
            self.assertEqual(sheet["M9"].value, 240)
            self.assertEqual(
                sum(1 for row in sheet.iter_rows() if row[0].value == "Cód:"),
                1,
            )
            workbook.close()
