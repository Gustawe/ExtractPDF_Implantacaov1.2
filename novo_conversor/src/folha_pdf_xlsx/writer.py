from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .audit_writer import write_audit_sheets
from .models import ConversionDetails, PayrollDocument


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBTLE_FILL = PatternFill("solid", fgColor="D9EAF7")
ERROR_FILL = PatternFill("solid", fgColor="F4CCCC")
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN_GRAY = Side(style="thin", color="D9E2F3")
MONEY_FORMAT = '#,##0.00;[Red]-#,##0.00'
DATE_FORMAT = "dd/mm/yyyy"


def write_workbook(
    documents: Iterable[PayrollDocument],
    output_path: str | Path,
) -> Path:
    docs = list(documents)
    if not docs:
        raise ValueError("Nenhum documento para exportar.")

    output = Path(output_path).resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)

    _write_employees(workbook, docs)
    _write_events(workbook, docs)
    _write_departments(workbook, docs)
    _write_rubrics(workbook, docs)
    _write_fiscal(workbook, docs)
    _write_processing(workbook, docs)
    write_audit_sheets(workbook, _combined_details(docs))

    workbook.save(output)
    _verify_saved_workbook(output)
    return output


def _write_employees(workbook: Workbook, documents: list[PayrollDocument]) -> None:
    headers = [
        "Arquivo",
        "Empresa",
        "Competência",
        "Página",
        "Chave",
        "Tipo",
        "Matrícula",
        "Nome",
        "CPF",
        "Situação",
        "Admissão",
        "Vínculo",
        "Centro de Custo",
        "Departamento",
        "Horas Mês",
        "Código Cargo",
        "Cargo",
        "CBO",
        "Filial",
        "Salário",
        "ND",
        "NF",
        "Proventos",
        "Descontos",
        "Informativa",
        "Informativa Dedutora",
        "Líquido",
        "Base INSS",
        "Excedente INSS",
        "Base FGTS",
        "Valor FGTS",
        "Base IRRF",
        "Observações",
    ]
    rows: list[list[Any]] = []
    row_by_employee: dict[tuple[str, str], int] = {}
    for document in documents:
        for employee in document.employees:
            row_by_employee[(employee.source_file, employee.employee_key)] = len(rows) + 2
            rows.append(
                [
                    employee.source_file,
                    f"{document.metadata.company_code} - {document.metadata.company_name}",
                    document.metadata.competence,
                    employee.page,
                    employee.employee_key,
                    employee.record_type,
                    employee.registration,
                    employee.name,
                    employee.cpf,
                    employee.status,
                    employee.admission_date,
                    employee.employment_type,
                    employee.cost_center,
                    employee.department,
                    employee.monthly_hours,
                    employee.job_code,
                    employee.job_title,
                    employee.cbo,
                    employee.branch,
                    employee.salary,
                    employee.dependents,
                    employee.family_dependents,
                    employee.total_earnings,
                    employee.total_discounts,
                    employee.informational,
                    employee.informational_deduction,
                    employee.net_amount,
                    employee.inss_base,
                    employee.inss_excess,
                    employee.fgts_base,
                    employee.fgts_value,
                    employee.irrf_base,
                    employee.observations,
                ]
            )
    sheet = _create_table_sheet(workbook, "Funcionarios", headers, rows, "FuncionariosTbl")
    target_columns = {
        "Soma de proventos": "W",
        "Soma de descontos": "X",
        "Cálculo do líquido": "AA",
    }
    for document in documents:
        for check in document.validations:
            row = row_by_employee.get((check.source_file, check.record_key))
            column = target_columns.get(check.check)
            if row and column:
                check.target_cell = f"Funcionarios!{column}{row}"
    _format_columns(sheet, money_columns=range(20, 21), date_columns=(11,))
    _format_columns(sheet, money_columns=range(23, 33))
    sheet.column_dimensions["H"].width = 34
    sheet.column_dimensions["Q"].width = 34
    sheet.column_dimensions["AG"].width = 52


def _write_events(workbook: Workbook, documents: list[PayrollDocument]) -> None:
    headers = [
        "Arquivo",
        "Página",
        "Chave Funcionário",
        "Matrícula",
        "Código",
        "Descrição",
        "Referência",
        "Valor",
        "Tipo",
        "Linha Original",
    ]
    rows = [
        [
            event.source_file,
            event.page,
            event.employee_key,
            event.registration,
            event.code,
            event.description,
            event.reference,
            event.value,
            event.kind,
            event.raw_text,
        ]
        for document in documents
        for event in document.events
    ]
    sheet = _create_table_sheet(workbook, "Verbas", headers, rows, "VerbasTbl")
    _format_columns(sheet, money_columns=(8,))
    sheet.column_dimensions["F"].width = 42
    sheet.column_dimensions["J"].width = 70


def _write_rubrics(workbook: Workbook, documents: list[PayrollDocument]) -> None:
    headers = [
        "Arquivo",
        "Página",
        "Código",
        "Descrição",
        "Referência",
        "Valor",
        "Tipo",
        "Linha Original",
    ]
    rows = [
        [
            item.source_file,
            item.page,
            item.code,
            item.description,
            item.reference,
            item.value,
            item.kind,
            item.raw_text,
        ]
        for document in documents
        for item in document.rubrics
    ]
    sheet = _create_table_sheet(
        workbook, "Resumo_Rubricas", headers, rows, "ResumoRubricasTbl"
    )
    _format_columns(sheet, money_columns=(6,))
    sheet.column_dimensions["D"].width = 44
    sheet.column_dimensions["H"].width = 70


def _write_departments(workbook: Workbook, documents: list[PayrollDocument]) -> None:
    headers = [
        "Arquivo",
        "Página",
        "Departamento",
        "Descrição",
        "Proventos",
        "Descontos",
        "Líquido",
        "Linha Original",
    ]
    rows = [
        [
            item.source_file,
            item.page,
            item.department,
            item.description,
            item.earnings,
            item.discounts,
            item.net_amount,
            item.raw_text,
        ]
        for document in documents
        for item in document.departments
    ]
    sheet = _create_table_sheet(
        workbook,
        "Resumo_Departamentos",
        headers,
        rows,
        "ResumoDepartamentosTbl",
    )
    _format_columns(sheet, money_columns=(5, 6, 7))
    sheet.column_dimensions["D"].width = 36
    sheet.column_dimensions["H"].width = 70


def _write_fiscal(workbook: Workbook, documents: list[PayrollDocument]) -> None:
    headers = [
        "Arquivo",
        "Página",
        "Seção",
        "Subgrupo",
        "Item",
        "Valor",
        "Linha Original",
    ]
    rows = [
        [
            item.source_file,
            item.page,
            item.section,
            item.subgroup,
            item.item,
            item.value,
            item.raw_text,
        ]
        for document in documents
        for item in document.fiscal_records
    ]
    sheet = _create_table_sheet(
        workbook, "Resumo_Fiscal", headers, rows, "ResumoFiscalTbl"
    )
    _format_columns(sheet, money_columns=(6,))
    for row_index in range(2, sheet.max_row + 1):
        if sheet.cell(row=row_index, column=3).value == "Situações":
            sheet.cell(row=row_index, column=6).number_format = "#,##0"
    sheet.column_dimensions["C"].width = 30
    sheet.column_dimensions["E"].width = 44
    sheet.column_dimensions["G"].width = 80


def _write_processing(workbook: Workbook, documents: list[PayrollDocument]) -> None:
    headers = [
        "Arquivo",
        "Empresa",
        "CNPJ",
        "Competência",
        "Perfil",
        "Páginas",
        "Empregados",
        "Contribuintes",
        "Verbas",
        "Rubricas",
        "Registros Fiscais",
        "Divergências",
        "Ocorrências",
        "Status",
        "Processado em",
    ]
    rows = []
    for document in documents:
        rows.append(
            [
                document.metadata.source_file,
                f"{document.metadata.company_code} - {document.metadata.company_name}",
                document.metadata.cnpj,
                document.metadata.competence,
                document.metadata.layout_profile,
                document.metadata.page_count,
                sum(1 for item in document.employees if item.record_type == "EMPREGADO"),
                sum(
                    1 for item in document.employees if item.record_type == "CONTRIBUINTE"
                ),
                len(document.events),
                len(document.rubrics),
                len(document.fiscal_records),
                sum(
                    1
                    for item in document.validations
                    if item.status in {"DIVERGÊNCIA", "FALHA"}
                ),
                len(document.issues),
                document.status,
                document.processed_at,
            ]
        )
    sheet = _create_table_sheet(
        workbook, "Processamento", headers, rows, "ProcessamentoTbl"
    )
    _format_columns(sheet, datetime_columns=(15,))
    sheet.column_dimensions["B"].width = 38
    sheet.column_dimensions["N"].width = 24
    if sheet.max_row >= 2:
        sheet.conditional_formatting.add(
            f"N2:N{sheet.max_row}",
            FormulaRule(formula=['N2="ERRO"'], fill=ERROR_FILL),
        )
        sheet.conditional_formatting.add(
            f"N2:N{sheet.max_row}",
            FormulaRule(formula=['N2="APROVADO COM DIVERGÊNCIAS"'], fill=WARNING_FILL),
        )
        sheet.conditional_formatting.add(
            f"N2:N{sheet.max_row}",
            FormulaRule(formula=['N2="APROVADO COM AVISOS"'], fill=WARNING_FILL),
        )


def _combined_details(documents: list[PayrollDocument]) -> ConversionDetails:
    return ConversionDetails(
        validations=[
            check for document in documents for check in document.validations
        ],
        issues=[issue for document in documents for issue in document.issues],
    )


def _create_table_sheet(
    workbook: Workbook,
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    table_name: str,
):
    sheet = workbook.create_sheet(title)
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.append(headers)
    for row in rows:
        sheet.append([_excel_value(value) for value in row])
    if not rows:
        sheet.append(["" for _ in headers])

    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 24

    last_column = get_column_letter(len(headers))
    table = Table(displayName=table_name, ref=f"A1:{last_column}{sheet.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.auto_filter.ref = f"A1:{last_column}{sheet.max_row}"
    _set_reasonable_widths(sheet)
    return sheet


def _excel_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    return value


def _set_reasonable_widths(sheet) -> None:
    for column_cells in sheet.iter_cols():
        letter = get_column_letter(column_cells[0].column)
        content_width = max(
            (len(str(cell.value)) for cell in column_cells if cell.value is not None),
            default=8,
        )
        sheet.column_dimensions[letter].width = min(max(content_width + 2, 11), 36)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top")


def _format_columns(
    sheet,
    money_columns: Iterable[int] = (),
    date_columns: Iterable[int] = (),
    datetime_columns: Iterable[int] = (),
) -> None:
    for column in money_columns:
        for cell in sheet.iter_cols(
            min_col=column, max_col=column, min_row=2, max_row=sheet.max_row
        ):
            for item in cell:
                item.number_format = MONEY_FORMAT
                item.alignment = Alignment(horizontal="right", vertical="top")
    for column in date_columns:
        for cell in sheet.iter_cols(
            min_col=column, max_col=column, min_row=2, max_row=sheet.max_row
        ):
            for item in cell:
                item.number_format = DATE_FORMAT
    for column in datetime_columns:
        for cell in sheet.iter_cols(
            min_col=column, max_col=column, min_row=2, max_row=sheet.max_row
        ):
            for item in cell:
                item.number_format = "dd/mm/yyyy hh:mm:ss"


def _verify_saved_workbook(path: Path) -> None:
    workbook = load_workbook(path, read_only=False, data_only=False)
    required_sheets = {
        "Funcionarios",
        "Verbas",
        "Resumo_Departamentos",
        "Resumo_Rubricas",
        "Resumo_Fiscal",
        "Processamento",
    }
    missing = required_sheets - set(workbook.sheetnames)
    if missing:
        raise ValueError(f"Planilha salva sem abas obrigatórias: {sorted(missing)}")
    if workbook["Processamento"].max_row < 2:
        raise ValueError("Aba Processamento está vazia.")
    workbook.close()
