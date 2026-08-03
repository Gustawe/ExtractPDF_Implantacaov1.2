from __future__ import annotations

import re
from calendar import monthrange
from datetime import date, datetime
from decimal import Decimal
from itertools import zip_longest
from pathlib import Path
from types import SimpleNamespace
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .audit_writer import write_audit_sheets
from .layout_models import LayoutLine, LayoutWord, PayrollLayoutDocument
from .models import EmployeeRecord, PayrollDocument, PayrollEvent, RubricSummary
from .parsing import parse_br_date, parse_br_decimal


NAVY = "17365D"
BLUE = "DCE6F1"
LIGHT_BLUE = "EAF2F8"
GOLD = "F4B183"
WHITE = "FFFFFF"
GRAY = "D9E1F2"
TEXT = "1F2937"
MONEY_FORMAT = '#,##0.00;[Red]-#,##0.00'
DATE_FORMAT = "dd/mm/yyyy"
THIN = Side(style="thin", color="B4C7DC")


def write_layout_workbook(
    document: PayrollLayoutDocument | PayrollDocument,
    output_path: str | Path,
) -> Path:
    if isinstance(document, PayrollDocument):
        return _write_structured_layout_workbook(document, output_path)

    output = Path(output_path).resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Folha"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.outlinePr.summaryBelow = True

    _configure_columns(sheet)
    row = 1
    for index, section in enumerate(document.sections, start=1):
        if index > 1:
            row += 2
        row = _write_section_header(sheet, row, section, index)
        for block in section.employee_blocks:
            start = row
            for line in block:
                values, role = _line_to_cells(line)
                _write_row(sheet, row, values, role)
                row += 1
            _style_employee_block(sheet, start, row - 1)
            sheet.row_dimensions.group(start, row - 1, outline_level=1, hidden=False)
            row += 1
        if section.summary_lines:
            row += 1
            for line in section.summary_lines:
                values, role = _line_to_cells(line, summary=True)
                _write_row(sheet, row, values, role)
                row += 1

    sheet.auto_filter.ref = f"A1:N{max(row - 1, 1)}"
    sheet.print_area = f"A1:N{max(row - 1, 1)}"
    sheet.sheet_properties.pageSetUpPr.autoPageBreaks = False
    write_audit_sheets(workbook, document.details)
    workbook.save(output)
    _verify(output, document)
    return output


def _write_structured_layout_workbook(
    document: PayrollDocument,
    output_path: str | Path,
) -> Path:
    """Write the parsed legacy report using the approved single-sheet layout."""
    output = Path(output_path).resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Folha"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.outlinePr.summaryBelow = True
    _configure_columns(sheet)

    groups = [
        (
            "Folha de Pagamento",
            [
                employee
                for employee in document.employees
                if employee.record_type != "CONTRIBUINTE"
            ],
        ),
        (
            "Folha de Pró-Labore",
            [
                employee
                for employee in document.employees
                if employee.record_type == "CONTRIBUINTE"
            ],
        ),
    ]

    row = 1
    sequence = 0
    for title, employees in groups:
        if not employees:
            continue
        sequence += 1
        if sequence > 1:
            row += 2
        section = _structured_section(document, title)
        row = _write_section_header(sheet, row, section, sequence)
        for employee in employees:
            start = row
            row = _write_structured_employee(sheet, row, employee)
            _map_employee_validation_cells(document, employee, row - 1)
            _style_employee_block(sheet, start, row - 1)
            sheet.row_dimensions.group(start, row - 1, outline_level=1, hidden=False)
            row += 1

    row = _write_structured_summaries(sheet, row, document)
    sheet.auto_filter.ref = f"A1:N{max(row - 1, 1)}"
    sheet.print_area = f"A1:N{max(row - 1, 1)}"
    sheet.sheet_properties.pageSetUpPr.autoPageBreaks = False
    write_audit_sheets(workbook, document.details)
    workbook.save(output)
    _verify(output, document)
    return output


def _structured_section(document: PayrollDocument, title: str) -> SimpleNamespace:
    period_start, period_end = _competence_period(document.metadata.competence)
    return SimpleNamespace(
        title=title,
        nickname=document.metadata.company_code,
        company_name=document.metadata.company_name,
        cnpj=document.metadata.cnpj,
        registration="",
        period_start=period_start,
        period_end=period_end,
        address="",
        district="",
        city="",
        state="",
    )


def _competence_period(competence: str) -> tuple[str, str]:
    match = re.fullmatch(r"(\d{2})/(\d{4})", competence.strip())
    if not match:
        return competence, competence
    month, year = (int(item) for item in match.groups())
    last_day = monthrange(year, month)[1]
    return f"01/{month:02d}/{year:04d}", f"{last_day:02d}/{month:02d}/{year:04d}"


def _write_structured_employee(
    sheet,
    row: int,
    employee: EmployeeRecord,
) -> int:
    registration: Any = (
        int(employee.registration)
        if employee.registration.isdigit()
        else employee.registration
    )
    header = [
        "Cód:",
        registration,
        "Nome:",
        "",
        employee.name,
        "",
        "",
        "",
        "Função:",
        employee.job_title,
        "",
        "Dep. IR:",
        employee.dependents if employee.dependents is not None else "",
        "",
    ]
    _write_row(sheet, row, header, "employee_header")
    row += 1

    admission: Any = (
        datetime.combine(employee.admission_date, datetime.min.time())
        if employee.admission_date
        else ""
    )
    meta = [
        "",
        "Admissão:",
        admission,
        "Situação:",
        employee.status,
        "",
        "",
        "",
        "Ocorrência:",
        "",
        "Salário:",
        _decimal_value(employee.salary),
        "",
        "",
    ]
    _write_row(sheet, row, meta, "employee_meta")
    row += 1

    earnings = [event for event in employee.events if event.kind == "P"]
    discounts = [event for event in employee.events if event.kind == "D"]
    for earning, discount in zip_longest(earnings, discounts):
        _write_row(
            sheet,
            row,
            _structured_event_row(earning, discount),
            "event",
        )
        row += 1

    base_rows = [
        [
            "",
            "Base INSS Empresa:",
            "",
            "",
            _decimal_value(employee.inss_base),
            "Base INSS Funcionário:",
            "",
            "",
            _decimal_value(employee.inss_base),
            "Base INSS Func. 13o. Salário:",
            "",
            "",
            "",
            "",
        ],
        [
            "",
            "Base F.G.T.S. 13o.:",
            "",
            "",
            "",
            "Base F.G.T.S.:",
            "",
            "",
            _decimal_value(employee.fgts_base),
            "F.G.T.S.:",
            "",
            "",
            _decimal_value(employee.fgts_value),
            "",
        ],
        [
            "",
            "Base I.R.R.F.:",
            "",
            "",
            _decimal_value(employee.irrf_base),
            "Deduções:",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ],
    ]
    for values in base_rows:
        _write_row(sheet, row, values, "base")
        row += 1

    totals = [
        "",
        "Proventos:",
        "",
        "",
        _decimal_value(employee.total_earnings),
        "Descontos:",
        "",
        "",
        _decimal_value(employee.total_discounts),
        "Liquido:",
        "",
        "",
        _decimal_value(employee.net_amount),
        "",
    ]
    _write_row(sheet, row, totals, "total")
    return row + 1


def _map_employee_validation_cells(
    document: PayrollDocument,
    employee: EmployeeRecord,
    totals_row: int,
) -> None:
    target_columns = {
        "Soma de proventos": "E",
        "Soma de descontos": "I",
        "Cálculo do líquido": "M",
    }
    for check in document.validations:
        if check.record_key != employee.employee_key:
            continue
        column = target_columns.get(check.check)
        if column:
            check.target_cell = f"Folha!{column}{totals_row}"


def _structured_event_row(
    earning: PayrollEvent | RubricSummary | None,
    discount: PayrollEvent | RubricSummary | None,
) -> list[Any]:
    values: list[Any] = [""] * 14
    if earning:
        values[1] = _code_value(earning.code)
        values[2] = earning.description
        values[5] = _number(earning.reference)
        values[6] = _decimal_value(earning.value)
    if discount:
        values[7] = _code_value(discount.code)
        values[8] = discount.description
        values[11] = _number(discount.reference)
        values[12] = _decimal_value(discount.value)
    return values


def _write_structured_summaries(
    sheet,
    row: int,
    document: PayrollDocument,
) -> int:
    if document.departments:
        row += 1
        row = _write_summary_band(sheet, row, "RESUMO POR DEPARTAMENTO")
        for item in document.departments:
            values = [
                item.department,
                item.description,
                "",
                "Proventos:",
                _decimal_value(item.earnings),
                "",
                "Descontos:",
                "",
                _decimal_value(item.discounts),
                "Liquido:",
                "",
                "",
                _decimal_value(item.net_amount),
                "",
            ]
            _write_row(sheet, row, values, "summary")
            row += 1

    if document.rubrics:
        row += 1
        row = _write_summary_band(sheet, row, "RESUMO POR RUBRICA")
        earnings = [item for item in document.rubrics if item.kind == "P"]
        discounts = [item for item in document.rubrics if item.kind == "D"]
        for earning, discount in zip_longest(earnings, discounts):
            _write_row(
                sheet,
                row,
                _structured_event_row(earning, discount),
                "event",
            )
            row += 1

    if document.fiscal_records:
        row += 1
        row = _write_summary_band(sheet, row, "RESUMO FISCAL")
        for item in document.fiscal_records:
            label = " — ".join(
                part for part in (item.section, item.subgroup, item.item) if part
            )
            values = ["", label, "", "", _decimal_value(item.value)]
            _write_row(sheet, row, values, "summary")
            row += 1
    return row


def _write_summary_band(sheet, row: int, title: str) -> int:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    cell = sheet.cell(row, 1, title)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(name="Aptos Display", size=12, bold=True, color=WHITE)
    cell.alignment = Alignment(horizontal="center")
    sheet.row_dimensions[row].height = 22
    return row + 1


def _code_value(value: str) -> int | str:
    return int(value) if value.isdigit() else value


def _decimal_value(value: Decimal | None) -> float | str:
    return float(value) if value is not None else ""


def _configure_columns(sheet) -> None:
    widths = [5, 13, 18, 12, 30, 16, 14, 8, 30, 17, 12, 16, 14, 14]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _write_section_header(sheet, row: int, section, sequence: int) -> int:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    cell = sheet.cell(row, 1, f"{section.title} — {section.period_start} a {section.period_end}")
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(name="Aptos Display", size=15, bold=True, color=WHITE)
    cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[row].height = 28
    row += 1

    rows = [
        ["", "Apelido:", section.nickname, "Razão Social:", "", section.company_name],
        ["", "CNPJ/CEI:", section.cnpj, "", "", "Inscrição:", section.registration,
         "", "", "", "Período de:", _date(section.period_start), "a", _date(section.period_end)],
        ["", "Endereço:", section.address, "", "", "", "", "Bairro:", section.district,
         "Cidade:", section.city, "UF:", section.state],
    ]
    for values in rows:
        _write_row(sheet, row, values, "header")
        row += 1
    row += 1
    return row


def _line_to_cells(
    line: LayoutLine, summary: bool = False
) -> tuple[list[Any], str]:
    text = line.text
    if text.startswith("Cód:"):
        return _employee_header(text), "employee_header"
    if text.startswith(("Admissão:", "Data Início:")):
        return _employee_status(text), "employee_meta"
    if text.startswith(("Base ", "Proventos:")):
        return _label_value_pairs(text), "total" if text.startswith("Proventos:") else "base"
    if text.startswith("R E S U M O"):
        return ["RESUMO"], "summary_title"
    if summary:
        if _looks_like_event(line):
            return _event_row(line), "event"
        return _summary_row(line), "summary"
    if _looks_like_event(line):
        return _event_row(line), "event"
    return _summary_row(line), "text"


def _employee_header(text: str) -> list[Any]:
    result: list[Any] = ["Cód:"]
    code = _capture(text, r"Cód:\s*(\d+)")
    name = _capture(text, r"Nome:\s*(.*?)(?:\s+Função:|\s+Dep\. IR:)")
    function = _capture(text, r"Função:\s*(.*?)\s+Dep\. IR:")
    dependent = _capture(text, r"Dep\. IR:\s*(\d+)")
    result.extend([int(code) if code else "", "Nome:", "", name, "", "", "", "Função:"])
    result.extend([function, "", "Dep. IR:", int(dependent) if dependent else "", ""])
    return result[:14]


def _employee_status(text: str) -> list[Any]:
    admission_label = "Data Início:" if text.startswith("Data Início:") else "Admissão:"
    admission = _capture(text, rf"{re.escape(admission_label)}\s*(\d{{2}}/\d{{2}}/\d{{4}})")
    status = _capture(text, r"Situação:\s*(.*?)(?:\s+Data:|\s+Ocorrência:|$)")
    event_date = _capture(text, r"Data:\s*(\d{2}/\d{2}/\d{4})")
    occurrence = _capture(text, r"Ocorrência:\s*(.*?)(?:\s+Salário:|$)")
    salary = _capture(text, r"Salário:\s*([\d.]+,\d{2})")
    values: list[Any] = ["", admission_label, _date(admission), "Situação:", status]
    if event_date:
        values.extend(["Data:", _date(event_date)])
    else:
        values.extend(["", ""])
    values.extend(["", "Ocorrência:", _integer(occurrence), "Salário:", _money(salary), ""])
    return values[:14]


def _event_row(line: LayoutLine) -> list[Any]:
    left = _parse_event_half(line.between(0, 300))
    right = _parse_event_half(line.between(300, 612))
    values: list[Any] = [""] * 14
    if left:
        values[1], values[2], values[5], values[6] = left
    if right:
        values[7], values[8], values[11], values[12] = right
    return values


def _parse_event_half(words: tuple[LayoutWord, ...]) -> tuple[Any, str, Any, Any] | None:
    if not words or not words[0].text.isdigit():
        return None
    code: Any = int(words[0].text)
    tokens = [word.text for word in words[1:]]
    numeric = [index for index, token in enumerate(tokens) if _money(token) is not None]
    if not numeric:
        return None
    value_index = numeric[-1]
    reference_index = numeric[-2] if len(numeric) > 1 else None
    end_description = reference_index if reference_index is not None else value_index
    description = " ".join(tokens[:end_description]).strip()
    reference = _number(tokens[reference_index]) if reference_index is not None else ""
    value = _money(tokens[value_index])
    return code, description, reference, value


def _looks_like_event(line: LayoutLine) -> bool:
    halves = (line.between(0, 300), line.between(300, 612))
    return any(words and words[0].text.isdigit() for words in halves)


PAIR_PATTERN = re.compile(
    r"([^:]+:)\s*([\d.]+,\d{2}|)(?=\s+[A-Za-zÀ-ÿ. /0-9]+:|$)"
)


def _label_value_pairs(text: str) -> list[Any]:
    pairs = [(label.strip(), value.strip()) for label, value in PAIR_PATTERN.findall(text)]
    values: list[Any] = [""] * 14
    starts = (1, 5, 9)
    value_columns = (4, 8, 12)
    for index, (label, value) in enumerate(pairs[:3]):
        values[starts[index]] = label
        values[value_columns[index]] = _money(value) if value else ""
    return values


def _summary_row(line: LayoutLine) -> list[Any]:
    text = line.text
    if not text:
        return [""] * 14
    if ":" in text:
        pairs = _label_value_pairs(text)
        if any(value not in ("", None) for value in pairs):
            return pairs
    values: list[Any] = [""] * 14
    values[0] = text
    return values


def _write_row(sheet, row: int, values: list[Any], role: str) -> None:
    values = list(values) + [""] * (14 - len(values))
    for column, value in enumerate(values[:14], start=1):
        cell = sheet.cell(row, column, value)
        cell.font = Font(name="Aptos", size=9, color=TEXT)
        cell.alignment = Alignment(vertical="top")
        money_columns = {
            "event": (7, 13),
            "base": (5, 9, 13),
            "total": (5, 9, 13),
            "summary": (5, 9, 13),
            "employee_meta": (12,),
        }.get(role, ())
        if isinstance(value, (float, int, Decimal)) and column in money_columns:
            cell.number_format = MONEY_FORMAT
            cell.alignment = Alignment(horizontal="right", vertical="top")
        if isinstance(value, datetime):
            cell.number_format = DATE_FORMAT
    if role == "header":
        for cell in sheet[row]:
            cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        for column in (2, 4, 6, 8, 10, 12):
            sheet.cell(row, column).font = Font(name="Aptos", size=9, bold=True, color=NAVY)
    elif role == "employee_header":
        for cell in sheet[row]:
            cell.fill = PatternFill("solid", fgColor=BLUE)
            cell.border = Border(top=THIN)
        for column in (1, 3, 9, 12):
            sheet.cell(row, column).font = Font(name="Aptos", size=9, bold=True, color=NAVY)
    elif role == "total":
        for cell in sheet[row]:
            cell.fill = PatternFill("solid", fgColor=GOLD)
            cell.font = Font(name="Aptos", size=9, bold=True, color=TEXT)
            cell.border = Border(top=THIN, bottom=THIN)
    elif role == "base":
        for cell in sheet[row]:
            cell.fill = PatternFill("solid", fgColor="F7F9FC")
    elif role == "summary_title":
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
        cell = sheet.cell(row, 1)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Aptos Display", size=12, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center")
        sheet.row_dimensions[row].height = 22
    elif role == "summary":
        sheet.cell(row, 1).font = Font(name="Aptos", size=9, bold=True, color=NAVY)
    sheet.row_dimensions[row].height = 15


def _style_employee_block(sheet, start: int, end: int) -> None:
    for row in range(start, end + 1):
        sheet.cell(row, 1).border = Border(left=THIN)
        sheet.cell(row, 14).border = Border(right=THIN)
    for cell in sheet[end]:
        cell.border = Border(
            left=cell.border.left,
            right=cell.border.right,
            bottom=THIN,
        )


def _capture(text: str, pattern: str) -> str:
    match = re.search(pattern, text)
    return match.group(1).strip() if match else ""


def _money(value: str) -> float | None:
    parsed = parse_br_decimal(value)
    return float(parsed) if parsed is not None else None


def _number(value: str) -> float | int | str:
    parsed = parse_br_decimal(value)
    if parsed is None:
        return value
    return int(parsed) if parsed == parsed.to_integral() else float(parsed)


def _integer(value: str) -> int | str:
    return int(value) if value.isdigit() else value


def _date(value: str) -> datetime | str:
    parsed = parse_br_date(value)
    return datetime.combine(parsed, datetime.min.time()) if parsed else value


def _verify(
    path: Path,
    document: PayrollLayoutDocument | PayrollDocument,
) -> None:
    workbook = load_workbook(path, read_only=False, data_only=False)
    if not workbook.sheetnames or workbook.sheetnames[0] != "Folha":
        raise ValueError("A exportação visual deve manter 'Folha' como aba principal.")
    sheet = workbook["Folha"]
    found = sum(
        1 for row in sheet.iter_rows() if row[0].value == "Cód:"
    )
    expected = (
        document.employee_count
        if isinstance(document, PayrollLayoutDocument)
        else len(document.employees)
    )
    if found != expected:
        raise ValueError(
            f"Quantidade de blocos divergente: esperado {expected}, salvo {found}."
        )
    workbook.close()
