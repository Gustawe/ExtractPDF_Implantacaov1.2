from __future__ import annotations

from decimal import Decimal

from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from .models import ConversionDetails, ValidationCheck


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
DIVERGENCE_FILL = PatternFill("solid", fgColor="FFF2CC")
WARNING_FILL = PatternFill("solid", fgColor="FCE4D6")
NOT_APPLICABLE_FILL = PatternFill("solid", fgColor="E7E6E6")
ERROR_FILL = PatternFill("solid", fgColor="F4CCCC")
MONEY_FORMAT = '#,##0.00;[Red]-#,##0.00'
TRACE_MONEY_FORMAT = '#,##0.00;-#,##0.00'


def write_audit_sheets(workbook, details: ConversionDetails) -> None:
    visible_validations = [
        check for check in details.validations if check.status != "OK"
    ]
    if visible_validations:
        _write_validations(workbook, details.validations)
    if details.issues:
        _write_issues(workbook, details)
    _highlight_mapped_cells(workbook, details.validations)


def _write_validations(workbook, checks: list[ValidationCheck]) -> None:
    headers = [
        "Situação",
        "Funcionário",
        "Identificador",
        "Validação",
        "Esperado",
        "Apurado",
        "Diferença",
        "Página",
        "Mensagem",
        "Arquivo",
        "Escopo",
        "Célula",
    ]
    rows = [
        [
            check.status,
            check.employee_name,
            check.employee_id,
            check.check,
            _excel_value(check.expected),
            _excel_value(check.actual),
            _excel_value(check.difference),
            check.page,
            check.message,
            check.source_file,
            check.scope,
            check.target_cell,
        ]
        for check in checks
    ]
    sheet = _table_sheet(workbook, "Validações", headers, rows, "ValidacoesAuditoriaTbl")
    for column in (5, 6, 7):
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row, column).number_format = MONEY_FORMAT
    widths = (20, 30, 18, 28, 16, 16, 16, 10, 64, 32, 18, 20)
    _set_widths(sheet, widths)
    for row in range(2, sheet.max_row + 1):
        status = str(sheet.cell(row, 1).value or "")
        fill = _validation_fill(status)
        if fill:
            sheet.cell(row, 1).fill = fill
            sheet.cell(row, 1).font = Font(bold=True)
        if status == "NÃO APLICÁVEL":
            for column in (5, 6, 7):
                sheet.cell(row, column).number_format = TRACE_MONEY_FORMAT


def _write_issues(workbook, details: ConversionDetails) -> None:
    headers = [
        "Severidade",
        "Código",
        "Funcionário",
        "Página",
        "Mensagem",
        "Arquivo",
    ]
    rows = [
        [
            issue.severity,
            issue.code,
            issue.employee_name,
            issue.page,
            issue.message,
            issue.source_file,
        ]
        for issue in details.issues
    ]
    sheet = _table_sheet(workbook, "Ocorrências", headers, rows, "OcorrenciasAuditoriaTbl")
    _set_widths(sheet, (18, 34, 30, 10, 70, 34))
    for row in range(2, sheet.max_row + 1):
        severity = str(sheet.cell(row, 1).value or "")
        fill = ERROR_FILL if severity == "ERRO" else WARNING_FILL
        sheet.cell(row, 1).fill = fill
        sheet.cell(row, 1).font = Font(bold=True)


def _table_sheet(workbook, title: str, headers: list[str], rows: list[list], table_name: str):
    if title in workbook.sheetnames:
        del workbook[title]
    sheet = workbook.create_sheet(title)
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    table = Table(displayName=table_name, ref=f"A1:{sheet.cell(1, len(headers)).column_letter}{sheet.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.auto_filter.ref = table.ref
    return sheet


def _set_widths(sheet, widths: tuple[int, ...]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width


def _validation_fill(status: str):
    if status in {"DIVERGÊNCIA", "FALHA"}:
        return DIVERGENCE_FILL
    if status == "AVISO":
        return WARNING_FILL
    if status == "NÃO APLICÁVEL":
        return NOT_APPLICABLE_FILL
    return None


def _highlight_mapped_cells(workbook, checks: list[ValidationCheck]) -> None:
    for check in checks:
        if check.status not in {"DIVERGÊNCIA", "FALHA"} or not check.target_cell:
            continue
        sheet_name, separator, coordinate = check.target_cell.rpartition("!")
        sheet_name = sheet_name.strip("'")
        if not separator or sheet_name not in workbook.sheetnames:
            continue
        try:
            cell = workbook[sheet_name][coordinate]
        except (KeyError, ValueError):
            continue
        cell.fill = DIVERGENCE_FILL
        cell.comment = Comment(_comment_text(check), "Conversor de Folhas")


def _comment_text(check: ValidationCheck) -> str:
    page = str(check.page) if check.page is not None else "não informada"
    return "\n".join(
        (
            f"Validação: {check.check}",
            f"Esperado: {_display_value(check.expected)}",
            f"Apurado: {_display_value(check.actual)}",
            f"Diferença: {_display_value(check.difference)}",
            f"Página de origem: {page}",
        )
    )


def _excel_value(value):
    return float(value) if isinstance(value, Decimal) else value


def _display_value(value) -> str:
    if isinstance(value, Decimal):
        formatted = f"{value:,.2f}"
        return f"R$ {formatted.replace(',', 'X').replace('.', ',').replace('X', '.')}"
    return "—" if value is None else str(value)
